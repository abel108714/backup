from selenium import webdriver
from selenium.webdriver.support.select import Select
import selenium.webdriver.common.alert
import time
import os
from openpyxl import Workbook
import xlwings as xw
#getData("xxxxxxxx",3,1)
#getData("xxxxxxxx",3,2)
def getDataByWebPage():
   list = []
   new_text=driver.find_elements_by_xpath(("//table[@id='GridView1']/tbody/tr"))
   i = 0
   result = []
   wb = Workbook()
   ws = wb.active
   #ws['A' + str(i+1)]=Data[0]
   #xw.save("111.xlsm")

   # 定義你所要連接的檔案名稱
   wb = xw.Book("C:\\Users\\udev77\\Desktop\\111.xlsm")

   # 以Dict的形式存取你所需要的Sheet
   ws = {}



   for text in new_text:
      #print(text.text)
      list.append(text.text)
      #

      #ws['A1'] = getRowData("11304021",3)
      #print(list)
      
      
      RowData=list[i].split("\n")
      Data=RowData[0].split(" ")
      #ws['A' + str(i+1)] = Data[0]
      wb.sheets[0].range('A1').value  = Data[0]
      #wx.save("C:\\Users\\udev77\\Desktop\\111.xlsm")
      #ws.append(result)
      #result = text.text
      if i==5:
         break
      i=i+1
   #ws.append(list)
   #print(list)
   #xw.save("111.xlsm")
   # return result
   #




#記錄開始執行時間
start = time.time()

username = "OGB"
password = "OG140"

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('blink-settings=imagesEnabled=false')#不載入圖片，提升速度
chrome_options.add_argument('--headless')#不顯示視窗執行畫面
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--start-maximized')#若要顯示則顯示最大化
DriverPath="C:\\ChromeDriver\\chromedriver.exe"
driver = webdriver.Chrome(DriverPath,options=chrome_options)
driver.get("http://www.logistics.com.tw/TMP/login.aspx?ReturnUrl=%2FTMP%2FDefault.aspx")

#帳號
username_textbox = driver.find_element_by_id("Login1_UserName")
username_textbox.send_keys(username)

#密碼
password_textbox = driver.find_element_by_id("Login1_Password")
password_textbox.send_keys(password)

#登入
login_button = driver.find_element_by_id("Login1_LoginButton")
login_button.click()

TreeView =driver.find_element_by_id('ctl00_TreeView1n5')
TreeView.click()
#庫別
sel = driver.find_element_by_css_selector("select[id='ctl00_ContentPlaceHolder1_lstWare']")
Select(sel).select_by_visible_text('觀音物流中心')
#良品?
RadioButton = driver.find_element_by_id('ctl00_ContentPlaceHolder1_RadioButton1')
RadioButton.click()
#查詢
Button = driver.find_element_by_id('ctl00_ContentPlaceHolder1_Button1')
Button.click()

# list = []
# new_text=driver.find_elements_by_xpath(("//table[@id='GridView1']/tbody/tr"))
# for text in new_text:
#    list.append(text.text)

# print(list[1])
# print(list[1].split(" ")[3])
#print(getData("11111111",3))
#print(getData("11304007",3))#11304021
#print("----------------------------------")
#print(getRowData("11111111",3))
#print(getDataByWebPage())#11304021
#print("----------------------------------")
getDataByWebPage()
# NewLines=getDataByWebPage()
# FileName="123.txt"
# fp = open(FileName, "w",encoding="utf-8")
# print(NewLines)
# fp.writelines(NewLines)
# fp.close()

# wb = Workbook()
# ws = wb.active
# #ws['A1'] = getRowData("11304021",3)
# ws.append(getDataByWebPage())
# wb.save("sss.xlsx")



driver.quit()#關閉瀏覽器

#記錄結束執行時間
end = time.time()


print("耗時 " + str(end-start) + " 秒")




   # print(list[1])
   # print(list[1].split(" ")[FieldIndex])
   # return







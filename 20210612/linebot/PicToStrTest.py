# import pytesseract
from PIL import Image
import re
import Date as d
import pytesseract
# from PIL import Image
import re
import sys
import os
# from FileDataAccess import *
from openpyxl import Workbook
import xlwings as xw

# import xlrd
# from PIL import ImageFile
# import re

# def main():
#     pytesseract.pytesseract.tesseract_cmd = r"C:\\Users\\udev77\\AppData\\Local\\Tesseract-OCR\\tesseract.exe"
#     img = Image.open(r"C:\\Users\\udev77\Desktop\\11_0.jpg")
#     #img.show()
#     # # print(pytesseract.image_to_string(img, lang="chi_tra"))
#     # data = pytesseract.image_to_string(img, lang="chi_tra")
#     # dataList = re.split(r' ',data) # split the string

#     # # resultList = [str(i.strip()) for i in dataList if i != ''] # remove the '' str and convert str to int.
#     # # print(resultList)
#     # print(data)
#     # print(dataList)
#     # for i in range(len(dataList)):
#     #     print("dataList["+ str(i) +"]:"+str(dataList[i]))
#     #     s = '王瑲賢'
#     #     pos = s.find(dataList[i])
#     #     print(pos)
#     #     # if dataList[i] in s:   # 使用in運算子檢查
#     #     #     print('字串中有\'王\'')
#     #     # else:
#     #     #     print('字串中沒有\'王\'')
        

# if __name__ == "__main__":

#     main()

def delBlank(str):
    return re.sub(r"\s+", "", str)

def delSymbol(str,symbol):
    return re.sub(r"\s+", symbol, str)

def is_number(num):
    pattern = re.compile(r'^[-+]?[-0-9]\d*\.\d*|[-+]?\.?[0-9]\d*$')
    result = pattern.match(num)
    if result:
        return True
    else:
        return False

def isChinese(str):
    zhPattern = re.compile(u'[\u4e00-\u9fa5]+')
    match = zhPattern.search(str)
    if match:
        return True
    else:
        return False

def isDigit(str):
    regex = r"\d+"
    matches = re.finditer(regex, str, re.MULTILINE)
    for matchNum, match in enumerate(matches, start=1):
        # print ("Match {matchNum} was found at {start}-{end}: {match}".format(matchNum = matchNum, start = match.start(), end = match.end(), match = match.group()))
        if len(str) == match.end():
            return True
        else:
            return False

def checkSubstring(dataList='None', start='None', end='None', SubString='None'):
    i=start
    ss=""
    # print("dataList: " + str(dataList))
    # print("start: " + str(start))
    # print("end: " + str(end))
    # print("SubString: " + str(SubString))
    if end == 'None':
        end=len(dataList)
    #     print("new end: " + str(end))
    # print("[i]: " + str(i))
    # print("[end]: " + str(end))
    while i <= end:#len(dataList[i])-1
        # print(i)
        # print(dataList[i])
        temp=dataList[i].strip().replace(',','')
        ss=str(ss)+str(temp)
        # print("ss: "+str(ss))
        # print("SubString: "+str(SubString))
        if ss.find(SubString) !=-1:
            return True

        # if ss==SubString and i<-end:
        #     return True
        
            
        i=i+1
        # print(ss)
        # print(isDigit(ss))
        # print(ss.isdigit())
    print("error")
    return False

def getSubstringLastLoc(dataList='None', start='None', end='None', SubString='None'):
    i=start
    ss=""
    # print("dataList: " + str(dataList))
    # print("start: " + str(start))
    # print("end: " + str(end))
    # print("SubString: " + str(SubString))
    if end == 'None':
        end=len(dataList)
        # print("new end: " + str(end))

    while i <= end:#len(dataList[i])-1
        # print(i)
        # print(dataList[i])
        temp=dataList[i].strip().replace(',','')
        ss=str(ss)+str(temp)
        # print("ss: "+str(ss))
        # print("SubString: "+str(SubString))
        if ss.find(SubString) !=-1:
            return i

        # if ss==SubString and i<-end:
        #     return True
        
            
        i=i+1
        # print(ss)
        # print(isDigit(ss))
        # print(ss.isdigit())
    print("error")
    return -1


def search_value_in_column(ws, search_string, column="A"):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return column, None


def search_value_in_col_index(ws, search_string, col_idx=1):
    # print("ws: "+str(ws))  
    
    max_row=ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
    # print(max_row)
    for row in range(1, max_row + 1):
        # print("row= "+str(row))
        # print(ws.cells(row,col_idx).value)
        if ws.cells(row,col_idx).value == search_string:
            return col_idx, row
    return col_idx, None


def search_value_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row
# def main():
#     pytesseract.pytesseract.tesseract_cmd = r"C:\\Users\\udev77\\AppData\\Local\\Tesseract-OCR\\tesseract.exe"
#     img = Image.open(r"C:\\Users\\udev77\\Desktop\\11_0.jpg")
#     #img.show()
#     # print(pytesseract.image_to_string(img, lang="chi_tra"))
#     data = pytesseract.image_to_string(img, lang="chi_tra+eng")
#     print(data)
#     dataList = re.split(r'\n ',data) # split the string
#     print(dataList)
#     # print(len(dataList))
#     for i in range(len(dataList)-1):
#         # print(str(i))
#         if dataList[i] != "\n" and i>0:
#             # if dataList[i] in "\n\n":
#             # print(str(dataList[i]))#line.split(',')[1]str(i) + str("==>") + 
#             # print(str(dataList[i]).split('\n\n')[0])
#             print(str(dataList[i]).strip('\n'))
#             # print(str(dataList[i]).split('\n\n')[2])
#     # resultList = str([int(i.strip()) for i in dataList if i != '']) # remove the '' str and convert str to int.
#     # print(resultList)
def checkData(Datalist, data, SplitStr):
    SplitData=str(data).split(SplitStr)
    i=0
    for i in range(len(SplitData)):
    	if SplitData[i] != "":
            Datalist.append(SplitData[i].replace(".","").replace(",","").strip("[").strip("_"))
            if SplitData[i].isDigit():
                print(Datalist)
            else:
                Datalist.pop(i)


def main(argv):
    args = sys.argv[1:]
    # print(args[0])
    # print("PP")
    # print(args[1])
    # print(args[2])
    # args = sys.argv[1:]
    # print(args[1])
    pytesseract.pytesseract.tesseract_cmd = r"C:\\Users\\udev77\\AppData\\Local\\Tesseract-OCR\\tesseract.exe"

    
    #img = Image.open(r"C:\\Users\\udev77\Desktop\\11_0.jpg")
    #argv
    img = Image.open(args[0])
    # print(img)
    
    # img.show()
    Image.LOAD_TRUNCATED_IMAGES = True
    # r,g,b=im.split()
    # om=Image.merge('RGB',(b,g,r))
    # om.save('mydog.jpg')
    # print("PP3")
    # print(pytesseract.image_to_string(img, lang="chi_tra+eng"))
    data = pytesseract.image_to_string(img, lang="chi_tra")
    print(data)

    dataList=[]
    dataList = re.split(r' ',data) # split the string1617172996588.jpg
    resultList = [(i.strip('\\n\\n')) for i in dataList if i != ''] # remove the '' str and convert str to int.
    CheckedData=[]
    print(resultList)
    i=0
    for i in range(len(resultList)):
        checkData(CheckedData, resultList[i], "\n")
        print(CheckedData)

    # print(resultList)
    # print(dataList)

    print("check: " + str(checkSubstring(dataList=dataList, start=1, SubString="月目標")))
    if checkSubstring(dataList=dataList, start=1,  SubString="月目標") == True:
        # print("123 ")
        i=0
        j=0
        path="S:\\網通部\網通部資料暫存區\★網通部績效小幫手專用★\業績目標.xlsx"

        wb = Workbook()
        
        ws = wb.active
        wb = xw.Book(path)
        app = xw.apps.active 
        app.visible=False
        # print("456 ")
        # wb = xlrd.open_workbook(path)
        # sheet = wb.sheet_by_index(0)
        ws = wb.sheets[0]
        print("(len):"+str(len(dataList)))
        while i < len(dataList)-1:#len(dataList[i])-1
            # print(dataList)
            # print(len(dataList)-1)
            ss=dataList[i].strip().replace(',','')
            # newDataAccess=FileDataAccess(0,'2','-','S:\\網通部\網通部資料暫存區\★網通部績效小幫手專用★\網通部下個月業積目標.txt')
            #S:\\網通部\網通部資料暫存區\★網通部績效小幫手專用★\業績目標.xlsx
            #print(ws.cell(row=cell.row, column=2).value) #change column number for any cell value you want
            # wb.sheets[0].range('D' + str(i+1)).value  = Data[0]
            # wb.sheets[0].range('E' + str(i+1)).value  = Data[1]
            # wb.sheets[0].range('F' + str(i+1)).value  = Data[2]
            # wb.sheets[0].range('G' + str(i+1)).value  = Data[3]
            # wb.sheets[0].range('H' + str(i+1)).value  = Data[4]
            # wb.sheets[0].range('J' + str(i+1)).value  = Data[6]
            #ws.cell(row=2, column=2).value = 2

            # print(wb.sheets[0].range('A' + str(wb.sheets[0].cells.last_cell.row)).end('up').row)
            # print('vvvv: '+str(wb.sheets[0]))
            # print('vvvv: '+str(wb.sheets[0].cells(3,5).value))
            # print(ws[1])
            # for row_num in range(sheet.nrows):
            #     row_value = sheet.row_values(row_num)
            #     if row_value[1] == 3600000:
            #         print(row_value)
            # for row_num in range(sheet.nrows):
            #     row_value = sheet.row_values(row_num)
            #     if row_value[1] == 3600000:
            #         print(row_value)
            c=d.Date()
            #print(c.getDay(weeks=-4))
            # print(c.getMonth(weeks=-4))

            # if searchStr(c.getMonth(weeks=-4)+"月目標"):
            #     print("ok")
            SearchStr=str(c.getMonth(weeks=-4))+str("月目標")
            # print(SearchStr)
            # print(ws.row_max)
            col_idx, row = search_value_in_col_index(ws, SearchStr, 5)
            # print("col_idx : "+str(col_idx))
            # print("row : "+str(row))
            # if ss.find("\n")
            #     d=[]
            #     d = re.split(r'\n',ss) 
            loc=getSubstringLastLoc(dataList=dataList, start=1, SubString="月目標")+1
            print("=================================")
            print("(ss): " + str(ss))
            ss=ss.strip("_")
            ss=ss.strip("|")
            ss=ss.strip("“")
            ss=ss.strip("一")
            ss=ss.strip(".")
            ss=ss.strip("[")
            ss=ss.strip("\n")
            print("(loc): " + str(loc))
            print("(ss): " + str(ss))
            
            print("(i): " + str(i))
            print("=================================")
            if ss != "" and i != 0 and i > loc:

                # print(ss)
                # print(isDigit(ss))

                # if isDigit(ss):
                    # newDataAccess.setData(str(j),[ss])
                    # print("j: "+str(j)+"[ss]: "+str(ss))
                    # print("j: "+str(j)+"[ss]: "+str([ss]))
                    # print(j+3)
                    # print(col_idx)
                print("loc: " + str(loc))

                print("((ss)): " + str(ss))
                wb.sheets[0].cells(j+3,col_idx+1).value=str(ss)
                # wb.sheets[0].cell(row=j+3, column=2).value
                j=j+1

            i=i+1

            

        wb.save(path)
        print("已更新個人業績目標")
        wb.close()
        app.kill()
        
        os._exit(0)
        # i=0
        # for i in range(6):
        #     # print(i)
        #     print(newDataAccess.getData(0,i))
    # wb = Workbook()
    # ws = wb.active 
    # # 定義你所要連接的檔案名稱
    # path="S:\\網通部\網通部資料暫存區\★網通部績效小幫手專用★\業績目標.xlsx"
    # wb = xw.Book(path)
    # #找第二列

    # for row in ws.iter_rows(min_row=2, min_col=1, max_col=13, max_row=9):
    #     for cell in row:
    #         print(cell.value)
    # row_count=13
    # while (i <= row_count):
    #     for item in sheet.iter_cols():
    #     #for row in sheet.iter_cols():
    #         first_value = item[i]
    #         print (first_value.value)
    # i+=1
    # for row in ws.iter_rows("4月目標"):
    #     print("value1")
    #     for cell in row:
    #         if cell.value == "4月目標":
    #             print("value2")


if __name__ == "__main__":
    print("Pic1")
    print(sys.argv[0])
    print(sys.argv[1])
    print("Pic2")
    main(sys.argv)
    print("Pic3")

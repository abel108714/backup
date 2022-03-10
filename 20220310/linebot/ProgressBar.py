from tkinter import *
from tkinter.ttk import *
import time
from openpyxl import load_workbook
class ProgressBar:
    window = Tk()

    percent = StringVar()
    text = StringVar()

    bar = Progressbar(window,orient=HORIZONTAL,length=300)
    bar.pack(pady=10)

    percentLabel = Label(window,textvariable=percent).pack()
    taskLabel = Label(window,textvariable=text).pack()
    BarValue=0
    total = 100
    h=0
    def __init__(self,BarValue=None,total=None):
        i=1
        GB = 100
        download = 0
        speed = 1

        if BarValue==None:
            self.BarValue=0
        else:
            self.BarValue = BarValue

        if total==None:
            self.total=100
        else:
            self.total = total
        
        self.Progressbar()
        # self.bar['value']=self.updateProgressbar(1)
        # self.bar['value']+=50
        # self.bar['value']+=self.updateProgressbar(1)
        # self.window.update_idletasks()
        # self.window.mainloop()

    # def getWBIndexByWorkbook(wb)
    #     i = 1
    #     For i = 1 To WorkBooks.count
    #         If WorkBooks(i).Name = wb.Name Then
    #             getWBIndexByWorkbook = i
    #             Exit Function
    #         End If
    #     Next i
    #     getWBIndexByWorkbook = 0

        

    def updateProgressbar(self,value):
        self.BarValue = value

    def getProgressbar(self):
            return self.BarValue

    def refresh(self):
        self.window.destroy()
        self.__init__(self,self.BarValue,self.total)

    def Progressbar(self):
        # self.BarValue=value
        # self.bar['value']=value
        time.sleep(0.05)
        # print("updateProgressbar"+str(value))
        # print("111")
        # self.window.update_idletasks()
        # # self.window.mainloop()
        # print("222")
        # print("updateProgressbar")
        # time.sleep(0.05)
        # self.bar['value']+=(speed/100)*100
        # print("window.update_idletasks")
        # self.window.update_idletasks()
        # print("11")
        # while self.bar['value'] <=99:
            
        #     # if self.BarValue != 0:
        #     while self.bar['value'] <=99:
                # print(self.BarValue)
        time.sleep(1)
        
        # path = 'S:\\網通部\\◎資訊\\年度績效數字分析\\COPR2320210120000088202101200001.xlsx'
        # wb2 = load_workbook(path)
        # print (wb2['第 1 頁'].max_row)
        newBar=(self.getProgressbar()/self.total)*100
        self.bar['value']=newBar
        # record=(self.BarValue/self.total)*100
        print(self.total)
        print(self.BarValue)
        # self.BarValue=self.BarValue+1
        # self.h=self.h+1
        # self.updateProgressbar(self.h)
        self.window.update_idletasks()
        # while True:
        # if self.bar['value']!=record:
        #     self.window.mainloop()
        self.refresh()
        self.window.mainloop()

        # print("22")
        #     download+=speed
        #     time.sleep(0.5)
        #     print("while")
        #     self.updateProgressbar(download)
        #     print("while.")
            # button = Button(window,text="download",command=updateProgressbar(download)).pack()
            # self.window.mainloop()
        # self.window.mainloop()

    def checkBookDataRow():
        pass



# def main(argv):
#    # args = sys.argv[1:]
#    # print(argv[1])
#    # print(args[2])
#    # print(args[3])

if __name__ == "__main__":
#    main(sys.argv)
    pb=ProgressBar()
    print("123")
    pb.updateProgressbar(16)









































# import asyncio

# def hello_world(loop):
#     """A callback to print 'Hello World' and stop the event loop"""
#     print('Hello World')
#     loop.stop()

# loop = asyncio.get_event_loop()

# # Schedule a call to hello_world()
# loop.call_soon(hello_world, loop)

# # Blocking call interrupted by loop.stop()
# try:
#     loop.run_forever()
# finally:
#     loop.close()





# import asyncio
# from concurrent.futures.thread import ThreadPoolExecutor

# from selenium import webdriver

# executor = ThreadPoolExecutor(10)


# def scrape(url, *, loop):
#     loop.run_in_executor(executor, scraper, url)


# def scraper(url):
#     DriverPath="C:\\ChromeDriver\\chromedriver.exe"
#     driver = webdriver.Chrome(DriverPath)
#     driver.get(url)


# loop = asyncio.get_event_loop()
# for url in ["https://www.google.com.tw/"] * 2:
#     scrape(url, loop=loop)

# loop.run_until_complete(asyncio.gather(*asyncio.all_tasks(loop)))